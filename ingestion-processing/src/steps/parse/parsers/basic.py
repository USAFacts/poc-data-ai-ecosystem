"""Basic parser for common file formats without external services."""

import csv
import io
import json
import time
from typing import Any

from steps.parse.parsers.base import (
    Parser,
    ParserError,
    ParseResult,
    ParsedSection,
    ParsedTable,
    ParsedKeyValue,
)


class BasicParser(Parser):
    """Basic parser for common file formats.

    Handles simple formats without requiring external services:
    - CSV/TSV files
    - JSON files
    - Plain text files
    - Excel files (requires openpyxl)

    Use this as a fallback when Claude Vision or other services aren't available,
    or for simple structured data that doesn't need advanced parsing.
    """

    parser_type = "basic"
    supported_formats = ["csv", "tsv", "json", "txt", "xlsx", "xls", "pdf"]

    def parse(
        self,
        data: bytes,
        filename: str,
        mime_type: str | None = None,
    ) -> ParseResult:
        """Parse a document using basic built-in parsers.

        Args:
            data: Raw document bytes
            filename: Original filename
            mime_type: MIME type if known

        Returns:
            ParseResult with extracted content

        Raises:
            ParserError: If parsing fails
        """
        start_time = time.time()

        format = self.get_format_from_filename(filename)

        try:
            if format in ("csv", "tsv"):
                result = self._parse_csv(data, format)
            elif format == "json":
                result = self._parse_json(data)
            elif format == "txt":
                result = self._parse_text(data)
            elif format in ("xlsx", "xls"):
                result = self._parse_excel(data)
            elif format == "pdf":
                result = self._parse_pdf(data)
            else:
                raise ParserError(f"Unsupported format: {format}")

            result.parser = self.parser_type
            result.parser_version = "1.0"
            result.processing_time_ms = int((time.time() - start_time) * 1000)

            return result

        except ParserError:
            raise
        except Exception as e:
            raise ParserError(f"Basic parsing failed: {e}") from e

    def _parse_csv(self, data: bytes, format: str = "csv") -> ParseResult:
        """Parse CSV/TSV file."""
        delimiter = "\t" if format == "tsv" else ","

        try:
            # Try UTF-8 first, then fall back to latin-1
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("latin-1")

            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)

            if not rows:
                return ParseResult(
                    plain_text="",
                    warnings=["Empty CSV file"],
                )

            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            # Create markdown table
            markdown_lines = []
            if headers:
                markdown_lines.append("| " + " | ".join(headers) + " |")
                markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for row in data_rows[:100]:  # Limit for markdown
                markdown_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")

            if len(data_rows) > 100:
                markdown_lines.append(f"\n*... and {len(data_rows) - 100} more rows*")

            table = ParsedTable(
                headers=headers,
                rows=data_rows,
                markdown="\n".join(markdown_lines),
                page_number=1,  # CSV is a single "sheet"
            )

            return ParseResult(
                markdown="\n".join(markdown_lines),
                plain_text=text,
                tables=[table],
                sections=[ParsedSection(
                    type="table",
                    content="\n".join(markdown_lines),
                    page_number=1,
                )],
            )

        except Exception as e:
            raise ParserError(f"CSV parsing failed: {e}") from e

    def _parse_json(self, data: bytes) -> ParseResult:
        """Parse JSON file."""
        try:
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("latin-1")

            json_data = json.loads(text)

            # Extract key-value pairs from top level
            key_values: list[ParsedKeyValue] = []
            tables: list[ParsedTable] = []

            if isinstance(json_data, dict):
                for key, value in json_data.items():
                    if isinstance(value, (str, int, float, bool)) or value is None:
                        value_type = "string"
                        if isinstance(value, bool):
                            value_type = "boolean"
                        elif isinstance(value, (int, float)):
                            value_type = "number"
                        key_values.append(ParsedKeyValue(
                            key=key,
                            value=value,
                            value_type=value_type,
                        ))
                    elif isinstance(value, list) and value:
                        # Try to convert array to table
                        if isinstance(value[0], dict):
                            headers = list(value[0].keys())
                            rows = [[item.get(h) for h in headers] for item in value if isinstance(item, dict)]
                            tables.append(ParsedTable(
                                headers=headers,
                                rows=rows,
                                title=key,
                            ))

            elif isinstance(json_data, list) and json_data:
                if isinstance(json_data[0], dict):
                    headers = list(json_data[0].keys())
                    rows = [[item.get(h) for h in headers] for item in json_data if isinstance(item, dict)]
                    tables.append(ParsedTable(
                        headers=headers,
                        rows=rows,
                    ))

            # Create markdown representation
            markdown = f"```json\n{json.dumps(json_data, indent=2)[:5000]}\n```"
            if len(text) > 5000:
                markdown += "\n\n*... truncated*"

            return ParseResult(
                markdown=markdown,
                plain_text=text,
                key_value_pairs=key_values,
                tables=tables,
                sections=[ParsedSection(
                    type="metadata",
                    content=markdown,
                )],
            )

        except json.JSONDecodeError as e:
            raise ParserError(f"Invalid JSON: {e}") from e

    def _parse_text(self, data: bytes) -> ParseResult:
        """Parse plain text file."""
        try:
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("latin-1")

            # Simple section detection based on blank lines
            sections: list[ParsedSection] = []
            paragraphs = text.split("\n\n")

            for para in paragraphs:
                para = para.strip()
                if not para:
                    continue

                # Check if it looks like a header
                lines = para.split("\n")
                if len(lines) == 1 and len(para) < 100 and para.isupper():
                    sections.append(ParsedSection(
                        type="heading",
                        content=para,
                        level=1,
                    ))
                elif len(lines) == 1 and para.startswith("#"):
                    level = len(para) - len(para.lstrip("#"))
                    sections.append(ParsedSection(
                        type="heading",
                        content=para.lstrip("#").strip(),
                        level=level,
                    ))
                else:
                    sections.append(ParsedSection(
                        type="paragraph",
                        content=para,
                    ))

            return ParseResult(
                markdown=text,
                plain_text=text,
                sections=sections,
            )

        except Exception as e:
            raise ParserError(f"Text parsing failed: {e}") from e

    def _parse_excel(self, data: bytes) -> ParseResult:
        """Parse Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ParserError(
                "openpyxl package not installed. "
                "Install with: pip install openpyxl"
            )

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            tables: list[ParsedTable] = []
            sections: list[ParsedSection] = []
            all_markdown: list[str] = []

            for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]

                # Get all rows
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert None to empty string, handle various types
                    clean_row = [
                        str(cell) if cell is not None else ""
                        for cell in row
                    ]
                    if any(clean_row):  # Skip completely empty rows
                        rows.append(clean_row)

                if not rows:
                    continue

                # Assume first row is headers
                headers = rows[0] if rows else []
                data_rows = rows[1:] if len(rows) > 1 else []

                # Create markdown table
                markdown_lines = [f"## {sheet_name}\n"]
                if headers:
                    markdown_lines.append("| " + " | ".join(str(h) for h in headers) + " |")
                    markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for row in data_rows[:100]:
                    markdown_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")

                if len(data_rows) > 100:
                    markdown_lines.append(f"\n*... and {len(data_rows) - 100} more rows*")

                table = ParsedTable(
                    headers=headers,
                    rows=data_rows,
                    title=sheet_name,  # Sheet name for reference
                    page_number=sheet_idx,  # Sheet index (1-based)
                    sheet_name=sheet_name,  # Explicit sheet name
                    markdown="\n".join(markdown_lines),
                )
                tables.append(table)

                # Add sheet as a section with sheet name and index
                sections.append(ParsedSection(
                    type="heading",
                    content=sheet_name,
                    title=sheet_name,
                    level=2,
                    page_number=sheet_idx,  # Sheet index
                    sheet_name=sheet_name,
                ))
                sections.append(ParsedSection(
                    type="table",
                    content="\n".join(markdown_lines),
                    title=sheet_name,
                    page_number=sheet_idx,
                    sheet_name=sheet_name,
                ))

                all_markdown.extend(markdown_lines)
                all_markdown.append("")

            return ParseResult(
                markdown="\n".join(all_markdown),
                plain_text="\n".join(all_markdown),
                tables=tables,
                sections=sections,
                page_count=len(workbook.sheetnames),
            )

        except Exception as e:
            raise ParserError(f"Excel parsing failed: {e}") from e

    def _parse_pdf(self, data: bytes) -> ParseResult:
        """Parse PDF file using PyMuPDF (pymupdf).

        Extracts text via structured blocks and tables via find_tables().
        Significantly faster and lower memory than pdfplumber.
        """
        try:
            import pymupdf
        except ImportError:
            raise ParserError(
                "pymupdf package not installed. "
                "Install with: pip install pymupdf"
            )

        pymupdf.TOOLS.mupdf_display_warnings(False)
        pymupdf.TOOLS.mupdf_warnings(False)

        try:
            tables: list[ParsedTable] = []
            sections: list[ParsedSection] = []
            all_markdown: list[str] = []
            all_text: list[str] = []
            warnings: list[str] = []

            doc = pymupdf.open(stream=data, filetype="pdf")
            page_count = len(doc)

            for page_num_idx, page in enumerate(doc):
                page_num = page_num_idx + 1
                page_header = f"## Page {page_num}\n"
                all_markdown.append(page_header)

                sections.append(ParsedSection(
                    type="heading",
                    content=f"Page {page_num}",
                    level=2,
                    page_number=page_num,
                ))

                # Extract tables using PyMuPDF's find_tables()
                page_tables = page.find_tables()

                if page_tables.tables:
                    for table_idx, table in enumerate(page_tables.tables):
                        # Extract as pandas-like structure
                        table_data = table.extract()

                        if not table_data or not table_data[0]:
                            continue

                        # First row as headers
                        headers = [str(cell) if cell else "" for cell in table_data[0]]
                        data_rows = []

                        for row in table_data[1:]:
                            clean_row = [str(cell) if cell else "" for cell in row]
                            data_rows.append(clean_row)

                        # Create markdown table
                        markdown_lines = []
                        if headers:
                            markdown_lines.append("| " + " | ".join(headers) + " |")
                            markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

                        for row in data_rows[:100]:
                            escaped_row = [str(cell).replace("|", "\\|") for cell in row]
                            markdown_lines.append("| " + " | ".join(escaped_row) + " |")

                        if len(data_rows) > 100:
                            markdown_lines.append(f"\n*... and {len(data_rows) - 100} more rows*")

                        table_markdown = "\n".join(markdown_lines)

                        parsed_table = ParsedTable(
                            headers=headers,
                            rows=data_rows,
                            title=f"Table {table_idx + 1}",
                            page_number=page_num,
                            markdown=table_markdown,
                        )
                        tables.append(parsed_table)

                        sections.append(ParsedSection(
                            type="table",
                            content=table_markdown,
                            title=f"Table {table_idx + 1}",
                            page_number=page_num,
                        ))

                        all_markdown.append(table_markdown)
                        all_markdown.append("")

                # Extract text using structured blocks for better layout
                blocks = page.get_text("blocks")
                page_text_parts: list[str] = []

                for block in blocks:
                    # block = (x0, y0, x1, y1, text, block_no, block_type)
                    if block[6] == 0:  # text block (not image)
                        text = block[4].strip()
                        if text:
                            page_text_parts.append(text)

                page_text = "\n".join(page_text_parts)
                if page_text:
                    all_text.append(f"--- Page {page_num} ---\n{page_text}")

                    if not page_tables.tables:
                        all_markdown.append(page_text)
                        all_markdown.append("")

                        sections.append(ParsedSection(
                            type="paragraph",
                            content=page_text,
                            page_number=page_num,
                        ))

            doc.close()

            if not tables:
                warnings.append("No tables detected in PDF - extracted text only")

            return ParseResult(
                markdown="\n".join(all_markdown),
                plain_text="\n\n".join(all_text),
                tables=tables,
                sections=sections,
                page_count=page_count,
                warnings=warnings if warnings else None,
            )

        except Exception as e:
            raise ParserError(f"PDF parsing failed: {e}") from e
